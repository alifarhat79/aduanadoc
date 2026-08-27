import io
import csv
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab para PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.database import get_db
from app.models import Despacho, DespachoItem

router = APIRouter(prefix="/exportar", tags=["Exportación"])

def get_filtered_despachos_and_items(
    db: Session,
    q: Optional[str],
    estado: Optional[str],
    importador: Optional[str],
    propietario: Optional[str] = None,
    origen: Optional[str] = None,
    anio: Optional[str] = None
):
    query = db.query(Despacho)
    if q:
        search_pattern = f"%{q}%"
        filters = [
            Despacho.numero_despacho.ilike(search_pattern),
            Despacho.importador_nombre.ilike(search_pattern),
            Despacho.exportador_nombre.ilike(search_pattern),
            Despacho.propietario.ilike(search_pattern),
            Despacho.pais_origen.ilike(search_pattern),
            Despacho.bl.ilike(search_pattern)
        ]
        if q.strip().isdigit() and len(q.strip()) == 4:
            filters.append(func.strftime("%Y", Despacho.fecha_despacho) == q.strip())
        query = query.filter(or_(*filters))

    if estado:
        query = query.filter(Despacho.estado_procesamiento == estado)
    if importador:
        query = query.filter(Despacho.importador_nombre.ilike(f"%{importador}%"))
    if propietario:
        query = query.filter(Despacho.propietario.ilike(f"%{propietario}%"))
    if origen:
        query = query.filter(Despacho.pais_origen.ilike(f"%{origen}%"))
    if anio:
        query = query.filter(func.strftime("%Y", Despacho.fecha_despacho) == str(anio).strip())

    despachos = query.order_by(desc(Despacho.fecha_despacho), desc(Despacho.id)).all()
    despacho_ids = [d.id for d in despachos]
    items = db.query(DespachoItem).filter(DespachoItem.despacho_id.in_(despacho_ids)).all() if despacho_ids else []
    return despachos, items

# 1. EXCEL MULTI-HOJA
@router.get("/excel")
async def export_excel(
    q: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    importador: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    origen: Optional[str] = Query(None),
    anio: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    despachos, items = get_filtered_despachos_and_items(
        db=db, q=q, estado=estado, importador=importador, propietario=propietario, origen=origen, anio=anio
    )

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    # HOJA 1: DESPACHOS
    ws1 = wb.active
    ws1.title = "Despachos"
    ws1.views.sheetView[0].showGridLines = True

    headers1 = [
        "ID", "Dueño / Cliente", "Nº Despacho", "Fecha", "Aduana", "Régimen", "Canal",
        "Importador", "RUC/Doc", "Exportador/Proveedor", "País Origen",
        "Despachante", "Modalidad", "BL / AWB", "Contenedor",
        "FOB ($)", "Flete ($)", "Seguro ($)", "CIF ($)", "Moneda",
        "Total Impuestos (Gs)", "Peso Neto (Kg)", "Bultos", "Estado", "Archivo"
    ]
    ws1.append(headers1)

    for d in despachos:
        ws1.append([
            d.id,
            d.propietario or "Sin Asignar",
            d.numero_despacho or "",
            d.fecha_despacho.strftime("%Y-%m-%d") if d.fecha_despacho else "",
            d.aduana or "",
            d.regimen or "",
            d.canal or "",
            d.importador_nombre or "",
            d.importador_documento or "",
            d.exportador_nombre or "",
            d.pais_origen or "",
            d.despachante_nombre or "",
            d.modalidad_transporte or "",
            d.bl or d.awb or "",
            d.contenedor or "",
            d.valor_fob or 0.0,
            d.valor_flete or 0.0,
            d.valor_seguro or 0.0,
            d.valor_cif or 0.0,
            d.moneda or "USD",
            d.total_general or 0.0,
            d.peso_neto or 0.0,
            d.cantidad_bultos or 0.0,
            d.estado_procesamiento or "",
            d.nombre_archivo_original or ""
        ])

    # HOJA 2: ÍTEMS CON COLUMNA DE MARCA Y CÓDIGO EAN SEPARADOS
    ws2 = wb.create_sheet(title="Ítems y Mercancías")
    ws2.views.sheetView[0].showGridLines = True
    headers2 = [
        "ID Ítem", "ID Despacho", "Nº Despacho", "Dueño", "Ítem", "Subítem",
        "Posición Arancelaria / NCM", "CÓDIGO / EAN / SKU", "MARCA", "Descripción del Producto",
        "Cantidad", "Unidad", "FOB Unitario ($)", "FOB Total ($)", "Pág. Origen"
    ]
    ws2.append(headers2)

    desp_map = {d.id: d for d in despachos}
    for it in items:
        parent = desp_map.get(it.despacho_id)
        ws2.append([
            it.id,
            it.despacho_id,
            parent.numero_despacho if parent else "",
            parent.propietario if parent else "",
            it.numero_item,
            it.numero_subitem or "",
            it.codigo_ncm or "",
            it.codigo_producto or "",
            it.marca or "Sin Marca",
            it.descripcion or "",
            it.cantidad or 0.0,
            it.unidad or "UNIDAD",
            it.valor_unitario or 0.0,
            it.valor_total or 0.0,
            it.pagina_origen or 1
        ])

    # HOJA 3: RESUMEN
    ws3 = wb.create_sheet(title="Resumen Ejecutivo")
    ws3.views.sheetView[0].showGridLines = True
    ws3.append(["Métrica", "Valor"])
    
    total_fob = sum(d.valor_fob or 0.0 for d in despachos)
    total_cif = sum(d.valor_cif or 0.0 for d in despachos)
    total_imp = sum(d.total_general or 0.0 for d in despachos)
    total_kg = sum(d.peso_neto or 0.0 for d in despachos)

    resumen_data = [
        ["Total de Despachos Registrados", len(despachos)],
        ["Total de Ítems / Subítems", len(items)],
        ["Valor FOB Total (USD)", total_fob],
        ["Valor CIF Total (USD)", total_cif],
        ["Total Impuestos / Tributos (Gs)", total_imp],
        ["Peso Neto Total (Kg)", total_kg],
        ["Fecha de Generación del Reporte", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    for r in resumen_data:
        ws3.append(r)

    # Formatos de columnas
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"despachos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# 2. EXPORTAR HTML
@router.get("/html", response_class=HTMLResponse)
async def export_html(
    q: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    importador: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    origen: Optional[str] = Query(None),
    anio: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    despachos, items = get_filtered_despachos_and_items(
        db=db, q=q, estado=estado, importador=importador, propietario=propietario, origen=origen, anio=anio
    )
    
    total_fob = sum(d.valor_fob or 0.0 for d in despachos)
    total_cif = sum(d.valor_cif or 0.0 for d in despachos)
    total_imp = sum(d.total_general or 0.0 for d in despachos)

    html_content = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Reporte de Despachos Aduaneros</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
        <style>
            @media print {{ .no-print {{ display: none; }} }}
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 13px; background: #f8fafc; padding: 20px; }}
            .header-box {{ background: #1e40af; color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
            .table-sm th {{ background: #f1f5f9; }}
        </style>
    </head>
    <body>
        <div class="container-fluid">
            <div class="d-flex justify-content-between align-items-center mb-3 no-print">
                <button onclick="window.print()" class="btn btn-primary btn-sm">🖨️ Imprimir / Guardar como PDF</button>
                <span class="text-muted small">Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</span>
            </div>

            <div class="header-box">
                <h3 class="mb-1">Reporte Consolidado de Despachos Aduaneros</h3>
                <p class="mb-0 small">Total Registros: {len(despachos)} | FOB Total: ${total_fob:,.2f} | CIF Total: ${total_cif:,.2f} | Impuestos: Gs. {total_imp:,.0f}</p>
            </div>

            <div class="card mb-4 shadow-sm">
                <div class="card-header bg-white fw-bold">Listado de Despachos</div>
                <div class="table-responsive">
                    <table class="table table-bordered table-striped table-sm mb-0 align-middle">
                        <thead>
                            <tr>
                                <th>ID</th>
                                <th>Dueño</th>
                                <th>Nº Despacho</th>
                                <th>Fecha</th>
                                <th>Importador</th>
                                <th>RUC</th>
                                <th>Proveedor</th>
                                <th>FOB ($)</th>
                                <th>CIF ($)</th>
                                <th>Canal</th>
                                <th>Estado</th>
                            </tr>
                        </thead>
                        <tbody>
                            {"".join([f"<tr><td>{d.id}</td><td>{d.propietario or '-'}</td><td><strong>{d.numero_despacho or '-'}</strong></td><td>{d.fecha_despacho or '-'}</td><td>{d.importador_nombre or '-'}</td><td>{d.importador_documento or '-'}</td><td>{d.exportador_nombre or '-'}</td><td>${d.valor_fob or 0:,.2f}</td><td>${d.valor_cif or 0:,.2f}</td><td>{d.canal or '-'}</td><td>{d.estado_procesamiento}</td></tr>" for d in despachos])}
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="card shadow-sm">
                <div class="card-header bg-white fw-bold">Detalle de Mercancías / Ítems (con Marca y Código Separados)</div>
                <div class="table-responsive">
                    <table class="table table-bordered table-sm mb-0 align-middle">
                        <thead>
                            <tr>
                                <th>ID Desp.</th>
                                <th>Ítem</th>
                                <th>NCM</th>
                                <th>CÓDIGO / EAN</th>
                                <th class="text-primary">MARCA</th>
                                <th>Descripción</th>
                                <th>Cantidad</th>
                                <th>FOB Total ($)</th>
                            </tr>
                        </thead>
                        <tbody>
                            {"".join([f"<tr><td>{it.despacho_id}</td><td>{it.numero_item}.{it.numero_subitem or 1}</td><td><code>{it.codigo_ncm or '-'}</code></td><td><strong>{it.codigo_producto or '-'}</strong></td><td><strong class='text-primary'>{it.marca or 'Sin Marca'}</strong></td><td>{it.descripcion or '-'}</td><td>{it.cantidad or 0:,.2f}</td><td>${it.valor_total or 0:,.2f}</td></tr>" for it in items])}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)

# 3. EXPORTAR PDF
@router.get("/pdf")
async def export_pdf(
    q: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    importador: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    origen: Optional[str] = Query(None),
    anio: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    despachos, _ = get_filtered_despachos_and_items(
        db=db, q=q, estado=estado, importador=importador, propietario=propietario, origen=origen, anio=anio
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor("#1E40AF"),
        spaceAfter=10
    )

    elements.append(Paragraph("Reporte Ejecutivo de Despachos Aduaneros", title_style))
    elements.append(Paragraph(f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total Despachos: {len(despachos)}", styles['Normal']))
    elements.append(Spacer(1, 10))

    data = [["ID", "Dueño", "Nº Despacho", "Fecha", "Importador", "RUC", "Proveedor", "FOB ($)", "CIF ($)", "Estado"]]
    for d in despachos:
        data.append([
            str(d.id),
            (d.propietario or "-")[:15],
            d.numero_despacho or "-",
            d.fecha_despacho.strftime("%d/%m/%Y") if d.fecha_despacho else "-",
            (d.importador_nombre or "-")[:20],
            d.importador_documento or "-",
            (d.exportador_nombre or "-")[:18],
            f"${d.valor_fob:,.2f}" if d.valor_fob else "$0.00",
            f"${d.valor_cif:,.2f}" if d.valor_cif else "$0.00",
            d.estado_procesamiento or "-"
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")])
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    filename = f"reporte_despachos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# 4. EXPORTAR GOOGLE SHEETS (CSV UTF-8 Universal)
@router.get("/google-sheet")
async def export_google_sheet(
    q: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    importador: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    origen: Optional[str] = Query(None),
    anio: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Genera un archivo CSV estándar con codificación UTF-8 con BOM y delimitador por comas,
    listo para abrir directamente en Google Sheets o usar con =IMPORTDATA().
    """
    despachos, _ = get_filtered_despachos_and_items(
        db=db, q=q, estado=estado, importador=importador, propietario=propietario, origen=origen, anio=anio
    )
    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "ID", "Dueno_Cliente", "Numero_Despacho", "Fecha", "Aduana", "Regimen", "Canal",
        "Importador", "RUC", "Proveedor", "BL_AWB", "Contenedor", "FOB", "Flete", "Seguro", "CIF", "Moneda", "Impuestos", "Estado"
    ])

    for d in despachos:
        writer.writerow([
            d.id,
            d.propietario or "Sin Asignar",
            d.numero_despacho or "",
            d.fecha_despacho.strftime("%Y-%m-%d") if d.fecha_despacho else "",
            d.aduana or "",
            d.regimen or "",
            d.canal or "",
            d.importador_nombre or "",
            d.importador_documento or "",
            d.exportador_nombre or "",
            d.bl or d.awb or "",
            d.contenedor or "",
            d.valor_fob or 0.0,
            d.valor_flete or 0.0,
            d.valor_seguro or 0.0,
            d.valor_cif or 0.0,
            d.moneda or "USD",
            d.total_general or 0.0,
            d.estado_procesamiento or ""
        ])

    output.seek(0)
    filename = f"google_sheets_despachos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# 5. CSV CLÁSICO
@router.get("/csv")
async def export_csv(
    q: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    importador: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    origen: Optional[str] = Query(None),
    anio: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    return await export_google_sheet(
        q=q, estado=estado, importador=importador, propietario=propietario, origen=origen, anio=anio, db=db
    )
