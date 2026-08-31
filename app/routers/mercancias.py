from fastapi import APIRouter, Depends, Request, Query, Response
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, or_
from typing import Optional, List
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.database import get_db
from app.models import Despacho, DespachoItem, MarcaSubitemEtiqueta
from app.templates_config import templates
from app.services.normalizer import parse_date
from app.services.ncm_helper import get_ncm_info
import time
from app.services.subitem_cleaner import (
    get_brand_subitems,
    get_brand_hierarchical_subitems,
    invalidate_brand_cache
)

router = APIRouter(prefix="/mercancias", tags=["Mercancías"])

_DROPDOWN_CACHE = {
    "todas_marcas": [],
    "todos_ncms_info": [],
    "todos_propietarios": [],
    "last_updated": 0
}

def get_cached_dropdown_data(db: Session):
    now = time.time()
    if now - _DROPDOWN_CACHE["last_updated"] < 120 and _DROPDOWN_CACHE["todas_marcas"]:
        return (
            _DROPDOWN_CACHE["todas_marcas"],
            _DROPDOWN_CACHE["todos_ncms_info"],
            _DROPDOWN_CACHE["todos_propietarios"]
        )

    todas_marcas = [r[0] for r in db.query(DespachoItem.marca).distinct().filter(DespachoItem.marca.isnot(None), DespachoItem.marca != "").order_by(DespachoItem.marca).all()]
    
    raw_ncms = [r[0] for r in db.query(DespachoItem.codigo_ncm).distinct().filter(DespachoItem.codigo_ncm.isnot(None), DespachoItem.codigo_ncm != "").order_by(DespachoItem.codigo_ncm).all()]
    todos_ncms_info = []
    for raw_n in raw_ncms:
        info = get_ncm_info(raw_n)
        display_label = f"{info['codigo_corto']} - {info['rubro']}" if info['rubro'] != "Mercancías diversas" else raw_n
        todos_ncms_info.append({
            "codigo": raw_n,
            "display": display_label
        })

    todos_propietarios = [r[0] for r in db.query(Despacho.propietario).distinct().filter(Despacho.propietario.isnot(None), Despacho.propietario != "").order_by(Despacho.propietario).all()]

    _DROPDOWN_CACHE["todas_marcas"] = todas_marcas
    _DROPDOWN_CACHE["todos_ncms_info"] = todos_ncms_info
    _DROPDOWN_CACHE["todos_propietarios"] = todos_propietarios
    _DROPDOWN_CACHE["last_updated"] = now

    return todas_marcas, todos_ncms_info, todos_propietarios

def get_filtered_items_query(
    db: Session,
    q: Optional[str] = None,
    marca: Optional[str] = None,
    ncm: Optional[str] = None,
    propietario: Optional[str] = None,
    despacho_id: Optional[int] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    query = db.query(DespachoItem, Despacho).join(Despacho, DespachoItem.despacho_id == Despacho.id)

    if q:
        query = query.filter(
            or_(
                DespachoItem.descripcion.ilike(f"%{q}%"),
                DespachoItem.marca.ilike(f"%{q}%"),
                DespachoItem.codigo_producto.ilike(f"%{q}%"),
                DespachoItem.codigo_ncm.ilike(f"%{q}%"),
                Despacho.numero_despacho.ilike(f"%{q}%"),
                Despacho.importador_nombre.ilike(f"%{q}%"),
                Despacho.propietario.ilike(f"%{q}%")
            )
        )
    if marca:
        query = query.filter(DespachoItem.marca.ilike(marca.strip()))
    if ncm:
        query = query.filter(DespachoItem.codigo_ncm == ncm)
    if propietario:
        query = query.filter(Despacho.propietario == propietario)
    if despacho_id:
        query = query.filter(DespachoItem.despacho_id == despacho_id)

    if fecha_desde:
        d_desde = parse_date(fecha_desde)
        if d_desde:
            query = query.filter(Despacho.fecha_despacho >= d_desde)

    if fecha_hasta:
        d_hasta = parse_date(fecha_hasta)
        if d_hasta:
            query = query.filter(Despacho.fecha_despacho <= d_hasta)

    return query.order_by(desc(Despacho.fecha_despacho), desc(Despacho.id), DespachoItem.numero_item, DespachoItem.numero_subitem)

@router.get("", response_class=HTMLResponse)
async def mercancias_view(
    request: Request,
    db: Session = Depends(get_db),
    q: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    ncm: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=10, le=200)
):
    import math

    # Consulta base filtrada
    items_query = get_filtered_items_query(db, q, marca, ncm, propietario, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

    # Conteo total y páginas
    total_items = items_query.count()
    total_pages = max(1, math.ceil(total_items / per_page))
    current_page = min(page, total_pages)

    # Obtener solo la página solicitada (LIMIT / OFFSET)
    results = items_query.offset((current_page - 1) * per_page).limit(per_page).all()

    # Formatear lista de items con objeto despacho padre
    items_list = []
    for item, desp in results:
        item.despacho = desp
        items_list.append(item)

    # Agregados y métricas calculados sobre la consulta base (sin order by para máximo rendimiento)
    base_agg_query = get_filtered_items_query(db, q, marca, ncm, propietario, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta).order_by(None)

    # 1. Total marcas únicas en el filtro actual
    total_marcas_filtradas = (
        base_agg_query.with_entities(func.count(func.distinct(DespachoItem.marca)))
        .filter(DespachoItem.marca.isnot(None), DespachoItem.marca != "")
        .scalar() or 0
    )

    # 2. Total NCMs únicos en el filtro actual
    total_ncms_filtrados = (
        base_agg_query.with_entities(func.count(func.distinct(DespachoItem.codigo_ncm)))
        .filter(DespachoItem.codigo_ncm.isnot(None), DespachoItem.codigo_ncm != "")
        .scalar() or 0
    )

    # 3. NCM Destacado (el más frecuente en el conjunto actual)
    top_ncm_row = (
        base_agg_query.with_entities(DespachoItem.codigo_ncm, func.count(DespachoItem.id).label("cnt"))
        .filter(DespachoItem.codigo_ncm.isnot(None), DespachoItem.codigo_ncm != "")
        .group_by(DespachoItem.codigo_ncm)
        .order_by(desc("cnt"))
        .first()
    )
    top_ncm_info = None
    top_ncm_count = 0
    if top_ncm_row:
        top_ncm_info = get_ncm_info(top_ncm_row[0])
        top_ncm_count = top_ncm_row[1]

    # 4. Estadísticas para mini KPIs de Marcas (Top 15) con sus modelos/subítems
    marca_counts_raw = (
        db.query(DespachoItem.marca, func.count(DespachoItem.id).label("total_items"), func.sum(DespachoItem.valor_total).label("total_fob"))
        .filter(DespachoItem.marca.isnot(None), DespachoItem.marca != "")
        .group_by(DespachoItem.marca)
        .order_by(desc("total_items"))
        .limit(15)
        .all()
    )
    marca_counts = []
    for r in marca_counts_raw:
        familias = get_brand_hierarchical_subitems(db, r.marca)
        marca_counts.append({
            "marca": r.marca,
            "total_items": r.total_items,
            "total_fob": r.total_fob,
            "familias": familias
        })

    # 5. Estadísticas para mini KPIs de NCMs Destacados (Top 12 con traducción de rubro)
    top_ncms_rows = (
        db.query(DespachoItem.codigo_ncm, func.count(DespachoItem.id).label("cnt"))
        .filter(DespachoItem.codigo_ncm.isnot(None), DespachoItem.codigo_ncm != "")
        .group_by(DespachoItem.codigo_ncm)
        .order_by(desc("cnt"))
        .limit(12)
        .all()
    )
    ncm_kpis = []
    for n_row in top_ncms_rows:
        info = get_ncm_info(n_row[0])
        ncm_kpis.append({
            "codigo": n_row[0],
            "codigo_corto": info["codigo_corto"],
            "rubro": info["rubro"],
            "capitulo_nombre": info["capitulo_nombre"],
            "total_items": n_row[1],
            "display": f"{info['codigo_corto']} {info['rubro']}"
        })

    # 6. País de Origen Líder
    top_pais_row = (
        base_agg_query.with_entities(DespachoItem.pais_origen, func.count(DespachoItem.id).label("cnt"))
        .filter(DespachoItem.pais_origen.isnot(None), DespachoItem.pais_origen != "")
        .group_by(DespachoItem.pais_origen)
        .order_by(desc("cnt"))
        .first()
    )
    top_pais = top_pais_row[0] if top_pais_row else None

    # Listas para filtros dropdown en memoria caché
    todas_marcas, todos_ncms_info, todos_propietarios = get_cached_dropdown_data(db)

    return templates.TemplateResponse(
        request=request,
        name="mercancias.html",
        context={
            "items": items_list,
            "total_items": total_items,
            "total_marcas_filtradas": total_marcas_filtradas,
            "total_ncms_filtrados": total_ncms_filtrados,
            "top_ncm_info": top_ncm_info,
            "top_ncm_count": top_ncm_count,
            "top_pais": top_pais,
            "marca_kpis": marca_counts,
            "ncm_kpis": ncm_kpis,
            "todas_marcas": todas_marcas,
            "todos_ncms": todos_ncms_info,
            "todos_propietarios": todos_propietarios,
            "q": q or "",
            "selected_marca": marca or "",
            "selected_ncm": ncm or "",
            "selected_propietario": propietario or "",
            "selected_fecha_desde": fecha_desde or "",
            "selected_fecha_hasta": fecha_hasta or "",
            "page": current_page,
            "per_page": per_page,
            "total_pages": total_pages,
            "has_prev": current_page > 1,
            "has_next": current_page < total_pages,
            "prev_page": current_page - 1,
            "next_page": current_page + 1
        }
    )

@router.get("/api/marcas/{marca}/subitems")
async def get_marca_subitems_api(marca: str, db: Session = Depends(get_db)):
    """Retorna las familias y variantes jerárquicas de una marca con nombres limpios y conteo."""
    familias = get_brand_hierarchical_subitems(db, marca)
    return {
        "status": "success",
        "marca": marca,
        "familias": familias
    }

@router.post("/api/marcas/{marca}/subitems")
async def save_marca_subitems_api(marca: str, request: Request, db: Session = Depends(get_db)):
    """Guarda o actualiza las etiquetas y familias personalizadas de una marca."""
    data = await request.json()
    subitems = data.get("subitems", [])

    # Eliminar etiquetas existentes para esta marca y guardar la lista actualizada
    db.query(MarcaSubitemEtiqueta).filter(MarcaSubitemEtiqueta.marca.ilike(marca.strip())).delete(synchronize_session=False)

    for idx, item in enumerate(subitems):
        nombre = item.get("nombre_limpio", "").strip()
        patron = item.get("patron_busqueda", "").strip()
        familia = item.get("familia", "").strip() or None
        if nombre and patron:
            db.add(MarcaSubitemEtiqueta(
                marca=marca.strip(),
                familia=familia,
                nombre_limpio=nombre,
                patron_busqueda=patron,
                orden=idx
            ))

    db.commit()
    invalidate_brand_cache(marca)
    updated = get_brand_hierarchical_subitems(db, marca)
    return {
        "status": "success",
        "message": f"Etiquetas de {marca} actualizadas correctamente.",
        "marca": marca,
        "familias": updated
    }

@router.delete("/api/subitems/{subitem_id}")
async def delete_subitem_api(subitem_id: int, db: Session = Depends(get_db)):
    """Elimina una etiqueta personalizada."""
    record = db.query(MarcaSubitemEtiqueta).filter(MarcaSubitemEtiqueta.id == subitem_id).first()
    if record:
        marca = record.marca
        db.delete(record)
        db.commit()
        invalidate_brand_cache(marca)
        return {"status": "success", "message": "Etiqueta eliminada.", "marca": marca}
    return {"status": "error", "message": "Etiqueta no encontrada."}

@router.get("/exportar/excel")
async def export_mercancias_excel(
    db: Session = Depends(get_db),
    q: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    ncm: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None)
):
    results = get_filtered_items_query(db, q, marca, ncm, propietario, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mercancías Consolidadas"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    headers = [
        "ID Despacho", "Nº Despacho", "Fecha Despacho", "Dueño / Cliente", "Importador",
        "Ítem", "Subítem", "NCM / Posición", "MARCA", "CÓDIGO / EAN",
        "Descripción del Producto", "Cantidad", "Unidad", "FOB Unitario ($)", "FOB Total ($)",
        "IVA (%)", "Arancel (%)", "País Origen", "Pág. Origen"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ws.row_dimensions[1].height = 26

    for item, desp in results:
        fecha_str = desp.fecha_despacho.strftime("%d/%m/%Y") if desp.fecha_despacho else ""
        row = [
            desp.id,
            desp.numero_despacho or "",
            fecha_str,
            desp.propietario or "",
            desp.importador_nombre or "",
            item.numero_item,
            item.numero_subitem or "",
            item.codigo_ncm or "",
            item.marca or "Sin Marca",
            item.codigo_producto or "",
            item.descripcion or "",
            item.cantidad or 0.0,
            item.unidad or "UNIDAD",
            item.valor_unitario or 0.0,
            item.valor_total or 0.0,
            f"{item.tasa_iva * 100:.1f}%" if item.tasa_iva is not None else "",
            f"{item.tasa_arancel * 100:.1f}%" if item.tasa_arancel is not None else "",
            item.pais_origen or "",
            item.pagina_origen or 1
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            elif isinstance(cell.value, int):
                cell.alignment = center_align

    # Ajustar anchos
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"mercancias_consolidadas_{len(results)}_items.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/exportar/csv")
@router.get("/exportar/google-sheet")
async def export_mercancias_csv(
    db: Session = Depends(get_db),
    q: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    ncm: Optional[str] = Query(None),
    propietario: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None)
):
    results = get_filtered_items_query(db, q, marca, ncm, propietario, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta).all()

    output = io.StringIO()
    output.write("\ufeff")  # BOM UTF-8 para Excel y Google Sheets
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "ID_DESPACHO", "NUMERO_DESPACHO", "FECHA_DESPACHO", "DUENO_CLIENTE", "IMPORTADOR",
        "ITEM", "SUBITEM", "NCM_POSICION", "MARCA", "CODIGO_EAN",
        "DESCRIPCION", "CANTIDAD", "UNIDAD", "FOB_UNITARIO", "FOB_TOTAL",
        "IVA_PCT", "ARANCEL_PCT", "PAIS_ORIGEN", "PAGINA"
    ])

    for item, desp in results:
        fecha_str = desp.fecha_despacho.strftime("%d/%m/%Y") if desp.fecha_despacho else ""
        writer.writerow([
            desp.id,
            desp.numero_despacho or "",
            fecha_str,
            desp.propietario or "",
            desp.importador_nombre or "",
            item.numero_item,
            item.numero_subitem or "",
            item.codigo_ncm or "",
            item.marca or "Sin Marca",
            item.codigo_producto or "",
            item.descripcion or "",
            item.cantidad or 0.0,
            item.unidad or "UNIDAD",
            item.valor_unitario or 0.0,
            item.valor_total or 0.0,
            f"{item.tasa_iva * 100:.1f}%" if item.tasa_iva is not None else "",
            f"{item.tasa_arancel * 100:.1f}%" if item.tasa_arancel is not None else "",
            item.pais_origen or "",
            item.pagina_origen or 1
        ])

    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=mercancias_filtradas_{len(results)}_items.csv"}
    )
