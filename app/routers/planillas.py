import io
import math
from datetime import datetime, date
from typing import Optional, List
from fastapi import APIRouter, Depends, Request, HTTPException, Query, Response, status, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.database import get_db
from app.models import PlanillaValoracion, PlanillaItem, DespachoItem, Despacho
from app.templates_config import templates
from app.services.client_sheet_importer import ClientSheetImporter

router = APIRouter(tags=["Planillas de Valoración"])

class ItemPayload(BaseModel):
    id: Optional[int] = None
    orden: int = 1
    item_catalogo_id: Optional[int] = None
    cantidad: float = 1.0
    mercaderia: str
    marca: Optional[str] = None
    codigo_producto: Optional[str] = None
    precio_factura: float = 0.0
    precio_normal: float = 0.0
    precio_total: float = 0.0
    observacion: Optional[str] = None

class PlanillaSaveRequest(BaseModel):
    id: Optional[int] = None
    titulo: Optional[str] = "PLANILLA DE VALORACION"
    despacho_numero: Optional[str] = None
    factura_comercial: Optional[str] = None
    importador: Optional[str] = None
    propietario: Optional[str] = None
    fecha_emision: Optional[str] = None
    estado: Optional[str] = "BORRADOR"  # BORRADOR o FINIQUITADO
    observaciones: Optional[str] = None
    items: List[ItemPayload] = []


# 1. LISTADO DE PLANILLAS
@router.get("/planillas", response_class=HTMLResponse)
async def list_planillas_view(
    request: Request,
    q: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    importador: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=5, le=100),
    db: Session = Depends(get_db)
):
    query = db.query(PlanillaValoracion)
    if q:
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                PlanillaValoracion.despacho_numero.ilike(search),
                PlanillaValoracion.factura_comercial.ilike(search),
                PlanillaValoracion.importador.ilike(search),
                PlanillaValoracion.propietario.ilike(search)
            )
        )
    if estado:
        query = query.filter(PlanillaValoracion.estado == estado)
    if importador:
        query = query.filter(PlanillaValoracion.importador.ilike(f"%{importador.strip()}%"))

    query = query.order_by(desc(PlanillaValoracion.updated_at), desc(PlanillaValoracion.id))

    total_items = query.count()
    total_pages = max(1, math.ceil(total_items / per_page))
    current_page = min(page, total_pages)
    planillas = query.offset((current_page - 1) * per_page).limit(per_page).all()

    # Métricas para tarjetas
    total_planillas = db.query(func.count(PlanillaValoracion.id)).scalar() or 0
    total_finiquitadas = db.query(func.count(PlanillaValoracion.id)).filter(PlanillaValoracion.estado == "FINIQUITADO").scalar() or 0
    total_borradores = db.query(func.count(PlanillaValoracion.id)).filter(PlanillaValoracion.estado == "BORRADOR").scalar() or 0
    suma_total_valor = db.query(func.sum(PlanillaValoracion.total_general)).scalar() or 0.0

    importadores = [
        r[0] for r in db.query(PlanillaValoracion.importador).distinct().filter(
            PlanillaValoracion.importador.isnot(None), PlanillaValoracion.importador != ""
        ).order_by(PlanillaValoracion.importador).all()
    ]

    return templates.TemplateResponse(
        request=request,
        name="planillas/planillas_list.html",
        context={
            "planillas": planillas,
            "q": q,
            "estado_sel": estado,
            "importador_sel": importador,
            "importadores": importadores,
            "total_planillas": total_planillas,
            "total_finiquitadas": total_finiquitadas,
            "total_borradores": total_borradores,
            "suma_total_valor": suma_total_valor,
            "page": current_page,
            "per_page": per_page,
            "total_pages": total_pages,
            "total_items": total_items,
            "has_prev": current_page > 1,
            "has_next": current_page < total_pages,
            "prev_page": current_page - 1,
            "next_page": current_page + 1
        }
    )


# 2. NUEVA PLANILLA (EDITOR TIPO EXCEL)
@router.get("/planillas/nueva", response_class=HTMLResponse)
async def nueva_planilla_view(request: Request, db: Session = Depends(get_db)):
    importadores_frecuentes = [
        r[0] for r in db.query(Despacho.importador_nombre).distinct().filter(
            Despacho.importador_nombre.isnot(None), Despacho.importador_nombre != ""
        ).order_by(Despacho.importador_nombre).all()
    ]
    return templates.TemplateResponse(
        request=request,
        name="planillas/planilla_editor.html",
        context={
            "planilla": None,
            "items": [],
            "fecha_hoy": date.today().strftime("%Y-%m-%d"),
            "importadores_frecuentes": importadores_frecuentes,
            "es_nueva": True
        }
    )


# 3. EDITAR / VER PLANILLA EXISTENTE
@router.get("/planillas/{planilla_id}", response_class=HTMLResponse)
async def editar_planilla_view(planilla_id: int, request: Request, db: Session = Depends(get_db)):
    planilla = db.query(PlanillaValoracion).filter(PlanillaValoracion.id == planilla_id).first()
    if not planilla:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")

    importadores_frecuentes = [
        r[0] for r in db.query(Despacho.importador_nombre).distinct().filter(
            Despacho.importador_nombre.isnot(None), Despacho.importador_nombre != ""
        ).order_by(Despacho.importador_nombre).all()
    ]

    return templates.TemplateResponse(
        request=request,
        name="planillas/planilla_editor.html",
        context={
            "planilla": planilla,
            "items": planilla.items,
            "fecha_hoy": planilla.fecha_emision.strftime("%Y-%m-%d") if planilla.fecha_emision else date.today().strftime("%Y-%m-%d"),
            "importadores_frecuentes": importadores_frecuentes,
            "es_nueva": False
        }
    )


# 4. IMPRESIÓN DIRECTA FORMATEADA (CSS PRINT LIMPIO)
@router.get("/planillas/{planilla_id}/imprimir", response_class=HTMLResponse)
async def imprimir_planilla_view(planilla_id: int, request: Request, db: Session = Depends(get_db)):
    planilla = db.query(PlanillaValoracion).filter(PlanillaValoracion.id == planilla_id).first()
    if not planilla:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")

    return templates.TemplateResponse(
        request=request,
        name="planillas/planilla_print.html",
        context={
            "planilla": planilla,
            "items": planilla.items
        }
    )


# 5. API BÚSQUEDA INTELIGENTE EN CATÁLOGO
@router.get("/api/planillas/buscar-catalogo")
async def buscar_catalogo_api(q: str = Query("", min_length=1), db: Session = Depends(get_db)):
    """Busca en vivo productos en el catálogo, ordenando el despacho más reciente al top y marcándolo."""
    term = f"%{q.strip()}%"
    results = (
        db.query(DespachoItem, Despacho)
        .join(Despacho, DespachoItem.despacho_id == Despacho.id)
        .filter(
            or_(
                DespachoItem.descripcion.ilike(term),
                DespachoItem.marca.ilike(term),
                DespachoItem.codigo_producto.ilike(term),
                DespachoItem.codigo_ncm.ilike(term)
            )
        )
        .order_by(desc(Despacho.fecha_despacho), desc(Despacho.id), desc(DespachoItem.id))
        .limit(30)
        .all()
    )

    items_data = []
    for idx, (item, despacho) in enumerate(results):
        unit_price = item.valor_unitario or 0.0
        if unit_price == 0.0 and item.cantidad and item.cantidad > 0 and item.valor_total:
            unit_price = round(item.valor_total / item.cantidad, 4)

        mercaderia_txt = item.descripcion or ""
        if item.marca and item.marca.upper() not in mercaderia_txt.upper():
            mercaderia_txt = f"{item.marca} - {mercaderia_txt}"

        items_data.append({
            "id": item.id,
            "mercaderia": mercaderia_txt,
            "marca": item.marca or "",
            "codigo_producto": item.codigo_producto or "",
            "codigo_ncm": item.codigo_ncm or "",
            "precio_normal": unit_price,
            "precio_factura": unit_price,
            "despacho_numero": despacho.numero_despacho or "",
            "importador": despacho.importador_nombre or "",
            "fecha": despacho.fecha_despacho.strftime("%d/%m/%Y") if despacho.fecha_despacho else "",
            "fecha_iso": despacho.fecha_despacho.isoformat() if despacho.fecha_despacho else "",
            "es_mas_reciente": (idx == 0)
        })

    return JSONResponse(content={"results": items_data})


# 5.1. API IMPORTAR ARCHIVO O PLANILLA DE CLIENTE (Excel, CSV, Word, PDF)
@router.post("/api/planillas/importar-archivo")
async def importar_archivo_cliente_api(file: UploadFile = File(...)):
    """
    Procesa un archivo subido por el cliente (Excel, CSV, Word, PDF) y retorna
    los ítems estructurados con marcas, cantidades y precios para insertar en la planilla.
    """
    try:
        content = await file.read()
        if not content:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "El archivo subido está vacío."}
            )

        res = ClientSheetImporter.import_file(content, file.filename or "archivo.xlsx")
        if not res.get("success"):
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": res.get("message", "No se pudieron extraer datos legibles del archivo.")}
            )

        return JSONResponse(content={
            "status": "success",
            "filename": file.filename,
            "total_items": res.get("total_items", 0),
            "headers": res.get("headers", []),
            "col_mapping": res.get("col_mapping", {}),
            "items": res.get("items", [])
        })
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Error al procesar el archivo: {str(e)}"}
        )


# 6. API GUARDAR O ACTUALIZAR PLANILLA (PERSISTENCIA MULTI-PC)
@router.post("/api/planillas")
async def guardar_planilla_api(payload: PlanillaSaveRequest, db: Session = Depends(get_db)):
    try:
        fecha_obj = date.today()
        if payload.fecha_emision:
            try:
                fecha_obj = datetime.strptime(payload.fecha_emision, "%Y-%m-%d").date()
            except Exception:
                pass

        if payload.id:
            planilla = db.query(PlanillaValoracion).filter(PlanillaValoracion.id == payload.id).first()
            if not planilla:
                raise HTTPException(status_code=404, detail="Planilla no encontrada")
        else:
            planilla = PlanillaValoracion()
            db.add(planilla)

        planilla.titulo = payload.titulo or "PLANILLA DE VALORACION"
        planilla.despacho_numero = (payload.despacho_numero or "").strip()
        planilla.factura_comercial = (payload.factura_comercial or "").strip()
        planilla.importador = (payload.importador or "").strip()
        planilla.propietario = (payload.propietario or "").strip()
        planilla.fecha_emision = fecha_obj
        planilla.estado = payload.estado or "BORRADOR"
        planilla.observaciones = payload.observaciones

        if payload.id:
            db.query(PlanillaItem).filter(PlanillaItem.planilla_id == planilla.id).delete()

        db.flush()

        tot_cantidad = 0.0
        tot_factura = 0.0
        tot_normal = 0.0
        tot_general = 0.0

        for idx, it in enumerate(payload.items, start=1):
            cant = float(it.cantidad or 0.0)
            p_fact = float(it.precio_factura or 0.0)
            p_norm = float(it.precio_normal or 0.0)
            
            p_calc = p_norm if p_norm > 0 else p_fact
            p_tot = float(it.precio_total or 0.0)
            if p_tot == 0.0:
                p_tot = round(cant * p_calc, 2)

            tot_cantidad += cant
            tot_factura += (cant * p_fact)
            tot_normal += (cant * p_norm)
            tot_general += p_tot

            new_item = PlanillaItem(
                planilla_id=planilla.id,
                orden=idx,
                item_catalogo_id=it.item_catalogo_id,
                cantidad=cant,
                mercaderia=it.mercaderia,
                marca=it.marca,
                codigo_producto=it.codigo_producto,
                precio_factura=p_fact,
                precio_normal=p_norm,
                precio_total=p_tot,
                observacion=it.observacion
            )
            db.add(new_item)

        planilla.total_cantidad = round(tot_cantidad, 2)
        planilla.total_precio_factura = round(tot_factura, 2)
        planilla.total_precio_normal = round(tot_normal, 2)
        planilla.total_general = round(tot_general, 2)

        db.commit()
        db.refresh(planilla)

        return JSONResponse(
            content={
                "success": True,
                "id": planilla.id,
                "estado": planilla.estado,
                "message": "Planilla guardada y finiquitada correctamente." if planilla.estado == "FINIQUITADO" else "Borrador de planilla guardado correctamente."
            }
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": f"Error al guardar la planilla: {str(e)}"}
        )


# 7. ELIMINAR PLANILLA
@router.delete("/api/planillas/{planilla_id}")
async def eliminar_planilla_api(planilla_id: int, db: Session = Depends(get_db)):
    planilla = db.query(PlanillaValoracion).filter(PlanillaValoracion.id == planilla_id).first()
    if not planilla:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")

    db.delete(planilla)
    db.commit()
    return JSONResponse(content={"success": True, "message": "Planilla eliminada exitosamente"})


# 8. EXPORTAR A EXCEL (.XLSX FORMATO IDÉNTICO)
@router.get("/planillas/{planilla_id}/excel")
async def exportar_planilla_excel(planilla_id: int, db: Session = Depends(get_db)):
    planilla = db.query(PlanillaValoracion).filter(PlanillaValoracion.id == planilla_id).first()
    if not planilla:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla de Valoracion"
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=18, bold=True, color="000000")
    font_sub_label = Font(name="Calibri", size=12, bold=True, underline="single", color="000000")
    font_sub_val = Font(name="Calibri", size=12, bold=True, color="1E40AF")
    font_header = Font(name="Calibri", size=11, bold=True, italic=True, color="000000")
    font_data = Font(name="Calibri", size=11, color="000000")
    font_total = Font(name="Calibri", size=12, bold=True, color="000000")

    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_total = PatternFill(start_color="E6EDF8", end_color="E6EDF8", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='thin', color='A0A0A0')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='double', color='000000')
    )

    ws.merge_cells("A1:E1")
    cell_title = ws["A1"]
    cell_title.value = (planilla.titulo or "PLANILLA DE VALORACION").upper()
    cell_title.font = font_title
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = "Despacho Nº:"
    ws["A2"].font = font_sub_label
    ws["A2"].alignment = Alignment(horizontal="right")
    ws.merge_cells("C2:E2")
    ws["C2"] = planilla.despacho_numero or "S/N"
    ws["C2"].font = font_sub_val

    ws.merge_cells("A3:B3")
    ws["A3"] = "FACTURA COMERCIAL NRO:"
    ws["A3"].font = font_sub_label
    ws["A3"].alignment = Alignment(horizontal="right")
    ws.merge_cells("C3:E3")
    ws["C3"] = planilla.factura_comercial or "-"
    ws["C3"].font = font_sub_val

    ws.merge_cells("A4:B4")
    ws["A4"] = "Importador:"
    ws["A4"].font = font_sub_label
    ws["A4"].alignment = Alignment(horizontal="right")
    ws.merge_cells("C4:E4")
    ws["C4"] = planilla.importador or "-"
    ws["C4"].font = font_sub_val

    ws.row_dimensions[5].height = 10

    headers = [
        ("A6", "Cantidad", Alignment(horizontal="center", vertical="center")),
        ("B6", "Mercaderias", Alignment(horizontal="center", vertical="center")),
        ("C6", "Precio\nFactura", Alignment(horizontal="center", vertical="center", wrap_text=True)),
        ("D6", "Precio\nNormal U$$", Alignment(horizontal="center", vertical="center", wrap_text=True)),
        ("E6", "Precio\nTotal U$$", Alignment(horizontal="center", vertical="center", wrap_text=True)),
    ]

    for pos, text, align in headers:
        cell = ws[pos]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = thin_border

    ws.row_dimensions[6].height = 28

    curr_row = 7
    for item in planilla.items:
        ws[f"A{curr_row}"] = item.cantidad or 0
        ws[f"A{curr_row}"].font = font_data
        ws[f"A{curr_row}"].alignment = Alignment(horizontal="center")
        ws[f"A{curr_row}"].border = thin_border
        ws[f"A{curr_row}"].number_format = "#,##0.00"

        ws[f"B{curr_row}"] = item.mercaderia or ""
        ws[f"B{curr_row}"].font = font_data
        ws[f"B{curr_row}"].alignment = Alignment(horizontal="left")
        ws[f"B{curr_row}"].border = thin_border

        ws[f"C{curr_row}"] = item.precio_factura or 0.0
        ws[f"C{curr_row}"].font = font_data
        ws[f"C{curr_row}"].alignment = Alignment(horizontal="right")
        ws[f"C{curr_row}"].border = thin_border
        ws[f"C{curr_row}"].number_format = "$#,##0.00"

        ws[f"D{curr_row}"] = item.precio_normal or 0.0
        ws[f"D{curr_row}"].font = font_data
        ws[f"D{curr_row}"].alignment = Alignment(horizontal="right")
        ws[f"D{curr_row}"].border = thin_border
        ws[f"D{curr_row}"].number_format = "$#,##0.00"

        ws[f"E{curr_row}"] = item.precio_total or (item.cantidad * item.precio_normal)
        ws[f"E{curr_row}"].font = font_data
        ws[f"E{curr_row}"].alignment = Alignment(horizontal="right")
        ws[f"E{curr_row}"].border = thin_border
        ws[f"E{curr_row}"].number_format = "$#,##0.00"

        curr_row += 1

    ws[f"A{curr_row}"] = planilla.total_cantidad or 0
    ws[f"A{curr_row}"].font = font_total
    ws[f"A{curr_row}"].fill = fill_total
    ws[f"A{curr_row}"].alignment = Alignment(horizontal="center")
    ws[f"A{curr_row}"].border = double_bottom_border
    ws[f"A{curr_row}"].number_format = "#,##0.00"

    ws[f"B{curr_row}"] = "TOTALES"
    ws[f"B{curr_row}"].font = font_total
    ws[f"B{curr_row}"].fill = fill_total
    ws[f"B{curr_row}"].alignment = Alignment(horizontal="center")
    ws[f"B{curr_row}"].border = double_bottom_border

    ws[f"C{curr_row}"] = planilla.total_precio_factura or 0.0
    ws[f"C{curr_row}"].font = font_total
    ws[f"C{curr_row}"].fill = fill_total
    ws[f"C{curr_row}"].alignment = Alignment(horizontal="right")
    ws[f"C{curr_row}"].border = double_bottom_border
    ws[f"C{curr_row}"].number_format = "$#,##0.00"

    ws[f"D{curr_row}"] = planilla.total_precio_normal or 0.0
    ws[f"D{curr_row}"].font = font_total
    ws[f"D{curr_row}"].fill = fill_total
    ws[f"D{curr_row}"].alignment = Alignment(horizontal="right")
    ws[f"D{curr_row}"].border = double_bottom_border
    ws[f"D{curr_row}"].number_format = "$#,##0.00"

    ws[f"E{curr_row}"] = planilla.total_general or 0.0
    ws[f"E{curr_row}"].font = font_total
    ws[f"E{curr_row}"].fill = fill_total
    ws[f"E{curr_row}"].alignment = Alignment(horizontal="right")
    ws[f"E{curr_row}"].border = double_bottom_border
    ws[f"E{curr_row}"].number_format = "$#,##0.00"

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Planilla_Valoracion_{planilla.despacho_numero or planilla.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# 9. EXPORTAR A PDF
@router.get("/planillas/{planilla_id}/pdf")
async def exportar_planilla_pdf(planilla_id: int, db: Session = Depends(get_db)):
    planilla = db.query(PlanillaValoracion).filter(PlanillaValoracion.id == planilla_id).first()
    if not planilla:
        raise HTTPException(status_code=404, detail="Planilla no encontrada")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=portrait(letter), rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        leading=20,
        alignment=1,
        textColor=colors.HexColor("#000000"),
        spaceAfter=15
    )

    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#1E293B")
    )

    elements.append(Paragraph("<b>PLANILLA DE VALORACION</b>", title_style))

    info_data = [
        [Paragraph(f"<b><u>Despacho Nº:</u></b> {planilla.despacho_numero or 'S/N'}", info_style), Paragraph(f"<b>Fecha:</b> {planilla.fecha_emision.strftime('%d/%m/%Y') if planilla.fecha_emision else '-'}", info_style)],
        [Paragraph(f"<b><u>FACTURA COMERCIAL NRO:</u></b> {planilla.factura_comercial or '-'}", info_style), Paragraph(f"<b>Estado:</b> {planilla.estado}", info_style)],
        [Paragraph(f"<b><u>Importador:</u></b> {planilla.importador or '-'}", info_style), Paragraph(f"<b>Dueño:</b> {planilla.propietario or '-'}", info_style)]
    ]
    info_table = Table(info_data, colWidths=[320, 220])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))

    table_data = [
        ["Cantidad", "Mercaderias", "Precio\nFactura", "Precio\nNormal U$$", "Precio\nTotal U$$"]
    ]

    for it in planilla.items:
        table_data.append([
            f"{it.cantidad:,.2f}",
            Paragraph(it.mercaderia or "", styles['Normal']),
            f"${it.precio_factura:,.2f}",
            f"${it.precio_normal:,.2f}",
            f"${it.precio_total:,.2f}"
        ])

    table_data.append([
        f"{planilla.total_cantidad:,.2f}",
        "TOTALES",
        f"${planilla.total_precio_factura:,.2f}",
        f"${planilla.total_precio_normal:,.2f}",
        f"${planilla.total_general:,.2f}"
    ])

    t = Table(table_data, colWidths=[70, 230, 80, 80, 80], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F1F5F9")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -2), 'LEFT'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('FONTSIZE', (0, 1), (-1, -2), 8.5),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#E2E8F0")),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)

    filename = f"Planilla_Valoracion_{planilla.despacho_numero or planilla.id}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
